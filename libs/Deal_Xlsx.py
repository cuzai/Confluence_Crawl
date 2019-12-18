import openpyxl
import os


class Deal_Xlsx:
    def __init__(self):
        cur_path = os.path.dirname(os.path.abspath(__file__))
        filename = os.path.join(cur_path, "template", "template.xlsx")
        # filename = "./template/template.xlsx"
        self.wb = openpyxl.load_workbook(filename)
        self.cur_sheet = None

    def set_cur_sheet(self, sheet_name):
        self.cur_sheet = self.wb[sheet_name]

    def write(self, col, row, val):
        sheet = self.cur_sheet
        sheet.cell(row=row, column=col, value=val)

    def merge_cell(self, fr, to):
        self.cur_sheet.merge_cells("{}:{}".format(fr, to))

    def change_font(self, cell, font_name, size, isbold):
        cell = self.cur_sheet[cell]
        cell.font = openpyxl.styles.Font(name=font_name, size=size, bold=isbold)

    def align(self, col, row, horizon, vertical, newline=False):
        cell = self.cur_sheet.cell(row=row, column=col)
        cell.alignment = openpyxl.styles.Alignment(
            horizontal=horizon, vertical=vertical, wrap_text=newline
        )

    def cell_color(self, cell, color):
        cell = self.cur_sheet[cell]
        cell.fill = openpyxl.styles.PatternFill(
            patternType="solid", fgColor=openpyxl.styles.Color(color)
        )

    def set_border(self, col, row):
        cell = self.cur_sheet.cell(row=row, column=col)
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(border_style="thin", color="FF000000"),
            right=openpyxl.styles.Side(border_style="thin", color="FF000000"),
            top=openpyxl.styles.Side(border_style="thin", color="FF000000"),
            bottom=openpyxl.styles.Side(border_style="thin", color="FF000000")
            #   diagonal=Side(border_style="thin",
            #   color='FF000000'),
            #   diagonal_direction=0,
            #   outline=Side(border_style="thin",
            #   color='FF000000'),
            #   vertical=Side(border_style="thin",
            #   color='FF000000'),
            #   horizontal=Side(border_style="thin",
            #   color='FF000000')
        )

    def set_newline(self, col, row):
        cell = self.cur_sheet.cell(row=row, column=col)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)


if __name__ == "__main__":
    dx = Deal_Xlsx()
    dx.set_cur_sheet("디지털사업부문")
    # print(dx.cur_sheet)

    dx.write(5, 5, "test")

    dx.wb.save("result.xlsx")

